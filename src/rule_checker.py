"""Brand, terminology, platform-rule, and quality prechecks for LocalizeFlow.

The checker is deterministic where fields and patterns are explicit. Checks that
depend on a live landing page, media file, legal context, or nuanced language
judgment require supplied evidence or human review. A quality score never
overrides fact or platform hard gates.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

from openpyxl import load_workbook

CHECKER_VERSION = "1.0.0"
ALLOWED_STATUSES = {"pass", "fail", "not_applicable", "needs_human_review"}
PLATFORM_NAMES = {
    "google_merchant_center": "Google Merchant Center",
    "tiktok_ads": "TikTok Ads",
    "generic_social": "Generic Social",
}
MARKET_PROFILES = {
    "US": {"language": "en-US", "currency": "USD"},
    "MX": {"language": "es-MX", "currency": "MXN"},
}
QUALITY_WEIGHTS = {
    "fact_accuracy": 30,
    "platform_fit": 20,
    "language_naturalness": 15,
    "brand_consistency": 15,
    "localization_quality": 10,
    "marketing_persuasiveness": 10,
}
SAFE_CTAS = {
    "en": {
        "explore the routine",
        "see product details",
        "build your simple routine",
        "learn how to use it",
    },
    "es": {
        "conoce la rutina",
        "consulta los detalles",
        "arma una rutina sencilla",
        "descubre cómo usarlo",
    },
}
HIGH_PRESSURE_CTA_PATTERNS = (
    r"\bbuy now before it(?:'|’)s gone\b",
    r"\btransform your skin today\b",
    r"\bget perfect skin\b",
    r"\bdoctors? recommend",
    r"\bcompra antes de que se agote\b",
    r"\btransforma tu piel hoy\b",
    r"\bconsigue una piel perfecta\b",
    r"\brecomendado por médicos\b",
)
PROMOTIONAL_PATTERNS = (
    r"\bfree shipping\b",
    r"\benvío gratis\b",
    r"\b\d+(?:\.\d{1,2})?\s*(?:USD|MXN)\b",
    r"[$€£]\s*\d",
    r"\b\d+%\s*(?:off|de descuento)\b",
    r"\b(?:sale|discount|coupon|promo code)\b",
)
URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)
HASHTAG_OR_AT_PATTERN = re.compile(r"(?:#\w+|@\w+|https?://\S+)", re.IGNORECASE)
ALL_CAPS_PATTERN = re.compile(r"\b[A-ZÁÉÍÓÚÜÑ]{4,}\b")
EXCESSIVE_PUNCTUATION_PATTERN = re.compile(r"([!?！？，。])\1+")
GIMMICKY_SYMBOL_PATTERN = re.compile(r"[★☆♥❤✓✔◆◇]{1,}")
ABNORMAL_SPACING_PATTERN = re.compile(r"[ \t]{2,}")
SPANISH_WORDS = {
    "para",
    "piel",
    "con",
    "sin",
    "una",
    "tu",
    "que",
    "los",
    "las",
    "del",
    "hidratante",
    "fragancia",
}
ENGLISH_WORDS = {
    "for",
    "skin",
    "with",
    "without",
    "your",
    "the",
    "and",
    "serum",
    "hydrating",
    "fragrance",
}
STOPWORDS = SPANISH_WORDS | ENGLISH_WORDS | {
    "a",
    "an",
    "to",
    "of",
    "in",
    "is",
    "it",
    "de",
    "el",
    "la",
    "y",
    "en",
    "se",
}


def _load_json(path: str | Path) -> dict[str, Any]:
    with Path(path).open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _normalize(text: Any) -> str:
    value = unicodedata.normalize("NFKC", str(text or "")).casefold()
    value = value.replace("–", "-").replace("—", "-").replace("’", "'")
    return re.sub(r"\s+", " ", value).strip()


def _strip_accents(text: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def _is_present(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, dict, tuple, set)):
        return bool(value)
    return True


def _get_path(data: dict[str, Any], path: str, default: Any = None) -> Any:
    current: Any = data
    for part in path.split("."):
        if not isinstance(current, dict) or part not in current:
            return default
        current = current[part]
    return current


def _valid_http_url(value: Any) -> bool:
    if not isinstance(value, str) or not value.strip():
        return False
    parsed = urlparse(value.strip())
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _language_family(language: str) -> str:
    return "es" if language.casefold().startswith("es") else "en"


def _term_pattern(term: str) -> re.Pattern[str]:
    escaped = re.escape(_normalize(term)).replace(r"\ ", r"\s+")
    if re.fullmatch(r"[\wÀ-ÿ' -]+", term, re.UNICODE):
        return re.compile(rf"(?<!\w){escaped}(?!\w)", re.IGNORECASE)
    return re.compile(escaped, re.IGNORECASE)


def _split_term_variants(raw: Any) -> list[str]:
    if not raw:
        return []
    cleaned = re.sub(r"（[^）]*）|\([^)]*\)", "", str(raw))
    values = []
    for item in re.split(r"\s*[|/]\s*", cleaned):
        item = item.strip()
        if item and item not in values:
            values.append(item)
    return values


def _content_segments(payload: dict[str, Any]) -> list[dict[str, str]]:
    content = payload.get("content") or {}
    segments: list[dict[str, str]] = []

    def add(location: str, value: Any) -> None:
        if isinstance(value, str) and value.strip():
            segments.append({"location": location, "text": value.strip()})

    add("content.title", content.get("title"))
    for index, bullet in enumerate(content.get("bullet_points") or []):
        if isinstance(bullet, dict):
            add(f"content.bullet_points[{index}]", bullet.get("text"))
        else:
            add(f"content.bullet_points[{index}]", bullet)
    add("content.description", content.get("description"))
    for index, scene in enumerate(content.get("scenes") or []):
        if isinstance(scene, dict):
            add(f"content.scenes[{index}].visual", scene.get("visual"))
            add(f"content.scenes[{index}].voiceover", scene.get("voiceover"))
            add(f"content.scenes[{index}].on_screen_text", scene.get("on_screen_text"))
    add("content.caption", content.get("caption"))
    add("content.hook", content.get("hook"))
    add("content.body", content.get("body"))
    add("content.cta", content.get("cta"))
    return segments


class RuleChecker:
    """Run rule checks without modifying the supplied content."""

    def __init__(
        self,
        platform_rules_path: str | Path,
        terminology_path: str | Path,
        prohibited_terms_path: str | Path,
    ) -> None:
        self.platform_rules_path = Path(platform_rules_path)
        self.terminology_path = Path(terminology_path)
        self.prohibited_terms_path = Path(prohibited_terms_path)
        self.platform_rules = _load_json(self.platform_rules_path)
        self.terminology = self._load_terminology(self.terminology_path)
        self.lexicon_rules = self._load_lexicon(self.prohibited_terms_path)

    @staticmethod
    def _load_terminology(path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(path, read_only=True, data_only=False)
        worksheet = workbook["Terminology"]
        headers = [cell.value for cell in worksheet[3]]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=4, values_only=True):
            row = dict(zip(headers, values))
            if row.get("term_id"):
                rows.append(row)
        workbook.close()
        return rows

    @staticmethod
    def _load_lexicon(path: Path) -> list[dict[str, str]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.DictReader(handle))

    def check(
        self,
        content_output: dict[str, Any],
        fact_check_output: dict[str, Any],
        previous_content_output: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        payload = deepcopy(content_output)
        fact_report = deepcopy(fact_check_output)
        segments = _content_segments(payload)
        platform_code = str(payload.get("platform") or "")
        platform = PLATFORM_NAMES.get(platform_code, platform_code)
        content_type = str(payload.get("content_type") or "")
        language = str(payload.get("language") or "")
        market = str(payload.get("market") or "")
        results: list[dict[str, Any]] = []

        applicable = [
            rule
            for rule in self.platform_rules.get("rules", [])
            if (
                rule.get("platform") == platform
                and rule.get("content_type") == content_type
            )
            or (
                rule.get("platform") == "Cross-platform"
                and rule.get("content_type") == "all_supported_content"
            )
        ]
        for rule in applicable:
            results.append(
                self._evaluate_platform_rule(rule, payload, fact_report, segments)
            )

        results.extend(self._check_lexicon(payload, segments))
        results.append(self._check_terminology(payload, segments))
        results.append(self._check_cta(payload))
        results.append(self._check_ai_disclosure(payload))
        results.append(self._check_brand_tone(payload, segments))

        summary = self._summarize(results)
        quality = self._score_quality(payload, fact_report, results, segments)
        export_gate = self._export_gate(fact_report, results, quality)
        human_review_items = [
            {
                "rule_id": item["rule_id"],
                "content_location": item["content_location"],
                "reason": item["reason"],
                "suggested_action": item["suggested_action"],
            }
            for item in results
            if item["status"] == "needs_human_review"
        ]
        suggestions = [
            {
                "priority": (
                    "P0"
                    if item["severity"] == "block"
                    else "P1"
                    if item["severity"] == "review"
                    else "P2"
                ),
                "rule_id": item["rule_id"],
                "location": item["content_location"],
                "reason": item["reason"],
                "suggested_action": item["suggested_action"],
            }
            for item in results
            if item["status"] in {"fail", "needs_human_review"}
        ]
        suggestions.sort(key=lambda item: (item["priority"], item["rule_id"]))

        version = payload.get("content_version") or {}
        current_version_id = version.get("version_id")
        previous_version_id = None
        if previous_content_output:
            previous_version_id = _get_path(
                previous_content_output, "content_version.version_id"
            )
        parent_version_id = version.get("parent_version_id")
        chain_valid = (
            not previous_content_output
            or parent_version_id == previous_version_id
        )
        version_record = {
            "content_id": version.get("content_id"),
            "current_version_id": current_version_id,
            "parent_version_id": parent_version_id,
            "previous_snapshot_version_id": previous_version_id,
            "chain_valid": chain_valid,
            "created_by": version.get("created_by"),
            "change_reason": version.get("change_reason"),
            "snapshots_are_immutable_copies": True,
            "current_content_snapshot": payload,
            "previous_content_snapshot": deepcopy(previous_content_output),
        }
        if previous_content_output and not chain_valid:
            export_gate = {
                **export_gate,
                "status": "blocked",
                "export_allowed": False,
                "blocking_priority": "content_version_integrity",
                "reason": (
                    "Content version parent does not match the supplied previous "
                    "snapshot; version history must be repaired before export."
                ),
            }

        return {
            "schema_version": "1.0",
            "checker_version": CHECKER_VERSION,
            "rule_check_id": f"rulecheck_{current_version_id or 'unknown'}",
            "checked_at": datetime.now(timezone.utc).isoformat(),
            "content_version_id": current_version_id,
            "sku": payload.get("sku"),
            "market": market,
            "language": language,
            "platform": platform_code,
            "content_type": content_type,
            "asset_versions": {
                "platform_rules": {
                    "path": str(self.platform_rules_path),
                    "rule_set_id": self.platform_rules.get("rule_set_id"),
                    "version": self.platform_rules.get("version"),
                    "verified_date": self.platform_rules.get("verified_date"),
                },
                "brand_voice": {
                    "path": "data/brand/brand_voice_guide.md",
                    "version": "1.0",
                },
                "terminology": {
                    "path": str(self.terminology_path),
                    "version": "1.0",
                    "record_count": len(self.terminology),
                },
                "prohibited_terms": {
                    "path": str(self.prohibited_terms_path),
                    "version": "1.0",
                    "record_count": len(self.lexicon_rules),
                },
                "fact_check_id": fact_report.get("fact_check_id"),
            },
            "hard_gate_snapshot": {
                "fact_gate": fact_report.get("export_gate"),
                "rule_gate": {
                    "block_failures": summary["totals"]["block_failures"],
                    "unresolved_hard_rules": summary["totals"][
                        "unresolved_hard_rules"
                    ],
                },
            },
            "rule_results": results,
            "summary_by_rule_type": summary,
            "quality_score": quality,
            "modification_suggestions": suggestions,
            "export_gate": export_gate,
            "human_review_items": human_review_items,
            "content_version_record": version_record,
            "limitations": [
                "Automated checks are prechecks, not platform approval or legal compliance.",
                "Live landing-page and media checks rely on supplied verification fields.",
                "Tone, naturalness, localization, and persuasion scores are heuristic and require human review.",
                "A quality score cannot override fact or platform hard-rule failures.",
            ],
        }

    def _evaluate_platform_rule(
        self,
        rule: dict[str, Any],
        payload: dict[str, Any],
        fact_report: dict[str, Any],
        segments: list[dict[str, str]],
    ) -> dict[str, Any]:
        rule_id = str(rule["rule_id"])
        content = payload.get("content") or {}
        platform_context = payload.get("platform_context") or {}
        feed = platform_context.get("feed_fields") or {}
        landing = platform_context.get("landing_page") or {}
        image = platform_context.get("image") or {}
        text = "\n".join(item["text"] for item in segments)
        title = str(content.get("title") or "")
        description = str(content.get("description") or "")
        claims = payload.get("claims") or []

        def result(
            status: str,
            reason: str,
            action: str,
            location: str = "",
            matched: Any = None,
            evidence: dict[str, Any] | None = None,
        ) -> dict[str, Any]:
            return self._result(
                rule_id=rule_id,
                rule_type=rule.get("rule_type", "project_internal_rule"),
                severity=rule.get("severity", "review"),
                status=status,
                content_location=location,
                matched_text=matched,
                reason=reason,
                source_url=rule.get("source_url"),
                suggested_action=action,
                evidence=evidence,
            )

        if rule_id == "GMC-H-001":
            identifier = feed.get("id")
            passed = isinstance(identifier, str) and 0 < len(identifier) <= 50
            return result(
                "pass" if passed else "fail",
                "Feed id is present and within 50 characters."
                if passed
                else "Google feed id is missing or exceeds 50 characters.",
                "Provide a unique, stable feed id of no more than 50 characters.",
                "platform_context.feed_fields.id",
                identifier,
            )

        if rule_id == "GMC-H-002":
            structured = feed.get("structured_title") or {}
            candidate = structured.get("content") or feed.get("title") or title
            passed = isinstance(candidate, str) and 0 < len(candidate) <= 150
            return result(
                "pass" if passed else "fail",
                f"Title length is {len(candidate) if isinstance(candidate, str) else 0}/150."
                if passed
                else "A title is required and must not exceed 150 characters.",
                "Add or shorten the title to 150 characters or fewer.",
                "content.title",
                candidate,
                {"character_count": len(candidate) if isinstance(candidate, str) else 0},
            )

        if rule_id == "GMC-H-003":
            ai_text = self._is_ai_text(payload)
            if not ai_text:
                return result(
                    "not_applicable",
                    "The title is not marked as generative-AI text.",
                    "No action required.",
                    "ai_disclosure.aigc_status",
                )
            structured = feed.get("structured_title") or {}
            passed = (
                _get_path(payload, "platform_fields.title_field_name")
                == "structured_title"
                and (
                    structured.get("digital_source_type")
                    or _get_path(payload, "platform_fields.digital_source_type")
                )
                == "trained_algorithmic_media"
            )
            return result(
                "pass" if passed else "fail",
                "AI title uses structured_title with trained_algorithmic_media."
                if passed
                else "Generative-AI title lacks the required structured_title disclosure.",
                "Use structured_title and set digital_source_type to trained_algorithmic_media.",
                "platform_fields",
            )

        if rule_id == "GMC-H-004":
            issues = self._editorial_issues(title, include_promotions=True)
            match = landing.get("title_matches_content")
            if issues:
                return result(
                    "fail",
                    f"Title contains prohibited editorial pattern(s): {', '.join(issues)}.",
                    "Remove promotional text, excessive capitalization, or gimmicky characters.",
                    "content.title",
                    title,
                    {"patterns": issues},
                )
            if match is not True:
                return result(
                    "needs_human_review",
                    "Title-to-landing-page consistency is not confirmed.",
                    "Confirm that the feed title accurately matches the landing page.",
                    "platform_context.landing_page.title_matches_content",
                    match,
                )
            return result(
                "pass",
                "Title editorial checks passed and landing-page title consistency is confirmed.",
                "No action required.",
                "content.title",
                title,
            )

        if rule_id == "GMC-H-005":
            structured = feed.get("structured_description") or {}
            candidate = (
                structured.get("content") or feed.get("description") or description
            )
            passed = isinstance(candidate, str) and 0 < len(candidate) <= 5000
            return result(
                "pass" if passed else "fail",
                f"Description length is {len(candidate) if isinstance(candidate, str) else 0}/5000."
                if passed
                else "A description is required and must not exceed 5000 characters.",
                "Add or shorten the description to 5000 characters or fewer.",
                "content.description",
                candidate,
                {"character_count": len(candidate) if isinstance(candidate, str) else 0},
            )

        if rule_id == "GMC-H-006":
            ai_text = self._is_ai_text(payload)
            if not ai_text:
                return result(
                    "not_applicable",
                    "The description is not marked as generative-AI text.",
                    "No action required.",
                    "ai_disclosure.aigc_status",
                )
            structured = feed.get("structured_description") or {}
            passed = (
                _get_path(payload, "platform_fields.description_field_name")
                == "structured_description"
                and (
                    structured.get("digital_source_type")
                    or _get_path(payload, "platform_fields.digital_source_type")
                )
                == "trained_algorithmic_media"
            )
            return result(
                "pass" if passed else "fail",
                "AI description uses structured_description with trained_algorithmic_media."
                if passed
                else "Generative-AI description lacks the required structured_description disclosure.",
                "Use structured_description and set digital_source_type to trained_algorithmic_media.",
                "platform_fields",
            )

        if rule_id == "GMC-H-007":
            issues = []
            if URL_PATTERN.search(description):
                issues.append("store_or_external_link")
            issues.extend(self._promotion_matches(description))
            match = landing.get("description_matches_content")
            if issues:
                return result(
                    "fail",
                    f"Description contains disallowed non-product or sales content: {', '.join(issues)}.",
                    "Remove links, promotions, pricing, competitor, or unrelated content.",
                    "content.description",
                    description,
                    {"patterns": issues},
                )
            if match is not True:
                return result(
                    "needs_human_review",
                    "Description-to-landing-page consistency is not confirmed.",
                    "Confirm that the description matches the landing page.",
                    "platform_context.landing_page.description_matches_content",
                    match,
                )
            return result(
                "pass",
                "Description contains product information only and landing-page consistency is confirmed.",
                "No action required.",
                "content.description",
            )

        if rule_id == "GMC-H-008":
            link = feed.get("link") or landing.get("url")
            if not _valid_http_url(link):
                return result(
                    "fail",
                    "A valid http/https landing-page link is required.",
                    "Provide a valid product landing-page URL.",
                    "platform_context.feed_fields.link",
                    link,
                )
            checks = {
                "domain_verified": landing.get("domain_verified"),
                "accessible": landing.get("accessible"),
            }
            if not all(value is True for value in checks.values()):
                return result(
                    "needs_human_review",
                    "Landing-page domain verification or accessibility is not confirmed.",
                    "Verify domain ownership and page accessibility in the target market.",
                    "platform_context.landing_page",
                    link,
                    checks,
                )
            return result(
                "pass",
                "Landing-page URL, domain verification, and accessibility are confirmed.",
                "No action required.",
                "platform_context.feed_fields.link",
                link,
            )

        if rule_id == "GMC-H-009":
            image_link = feed.get("image_link") or image.get("url")
            if not _valid_http_url(image_link):
                return result(
                    "fail",
                    "A valid main-image URL is required.",
                    "Provide a crawlable product image URL.",
                    "platform_context.feed_fields.image_link",
                    image_link,
                )
            checks = {
                "accurately_depicts_product": image.get("accurately_depicts_product"),
                "crawlable": image.get("crawlable"),
                "no_promotional_text": image.get("no_promotional_text"),
                "no_watermark": image.get("no_watermark"),
                "no_border": image.get("no_border"),
                "not_placeholder_or_generic": image.get("not_placeholder_or_generic"),
            }
            if not all(value is True for value in checks.values()):
                return result(
                    "needs_human_review",
                    "Main-image content and crawlability are not fully confirmed.",
                    "Review the image against all Google main-image requirements.",
                    "platform_context.image",
                    image_link,
                    checks,
                )
            return result(
                "pass",
                "Main-image URL and supplied image checks meet the rule.",
                "No action required.",
                "platform_context.image",
                image_link,
            )

        if rule_id == "GMC-H-010":
            if not self._is_ai_image(payload):
                return result(
                    "not_applicable",
                    "The product image is not marked as generative or significantly edited.",
                    "No action required.",
                    "platform_context.image.ai_origin",
                )
            metadata = image.get("iptc_digital_source_type")
            accepted = {"TrainedAlgorithmicMedia", "CompositeSynthetic", "AlgorithmicMedia"}
            passed = metadata in accepted
            return result(
                "pass" if passed else "fail",
                "Required IPTC digital-source metadata is present."
                if passed
                else "AI-generated product image lacks accepted IPTC source metadata.",
                "Preserve an accepted IPTC DigitalSourceType value in the product image.",
                "platform_context.image.iptc_digital_source_type",
                metadata,
            )

        if rule_id == "GMC-H-011":
            availability = feed.get("availability")
            allowed = {"in_stock", "out_of_stock", "preorder", "backorder"}
            if availability not in allowed:
                return result(
                    "fail",
                    "Availability is missing or not an allowed value.",
                    "Use an allowed availability value.",
                    "platform_context.feed_fields.availability",
                    availability,
                )
            if landing.get("availability_matches") is not True:
                return result(
                    "needs_human_review",
                    "Availability consistency with landing and checkout pages is not confirmed.",
                    "Confirm availability across feed, landing page, and checkout.",
                    "platform_context.landing_page.availability_matches",
                )
            return result(
                "pass",
                "Availability value is allowed and consistency is confirmed.",
                "No action required.",
                "platform_context.feed_fields.availability",
                availability,
            )

        if rule_id == "GMC-H-012":
            price = feed.get("price") or {}
            amount = price.get("amount") if isinstance(price, dict) else None
            currency = price.get("currency") if isinstance(price, dict) else None
            expected = MARKET_PROFILES.get(str(payload.get("market") or ""), {}).get(
                "currency"
            )
            valid_amount = isinstance(amount, (int, float)) and amount >= 0
            if not valid_amount or not currency or (expected and currency != expected):
                return result(
                    "fail",
                    "Price is missing, invalid, or uses the wrong market currency.",
                    f"Provide a non-negative price in {expected or 'the target-market currency'}.",
                    "platform_context.feed_fields.price",
                    price,
                    {"expected_currency": expected},
                )
            if landing.get("price_matches") is not True:
                return result(
                    "needs_human_review",
                    "Price consistency with landing and checkout pages is not confirmed.",
                    "Confirm amount and currency across feed, landing page, and checkout.",
                    "platform_context.landing_page.price_matches",
                    price,
                )
            return result(
                "pass",
                "Price and market currency are valid and consistency is confirmed.",
                "No action required.",
                "platform_context.feed_fields.price",
                price,
            )

        if rule_id == "GMC-H-013":
            fact_gate = _get_path(fact_report, "export_gate.status")
            prohibited_hits = self._matched_lexicon_rules(segments, "prohibited")
            if fact_gate == "blocked" or prohibited_hits:
                return result(
                    "fail",
                    "The prior fact gate is blocked or prohibited misleading claims were detected.",
                    "Resolve fact-check failures and remove unsupported or misleading claims.",
                    "hard_gate_snapshot.fact_gate",
                    [item["rule_id"] for item in prohibited_hits],
                    {"fact_gate": fact_gate},
                )
            return result(
                "pass",
                "No blocked fact-check result or prohibited misleading claim was detected.",
                "No action required.",
                "hard_gate_snapshot.fact_gate",
                fact_gate,
            )

        if rule_id == "GMC-H-014":
            required_value = landing.get("required_information")
            required: dict[str, Any] = (
                required_value if isinstance(required_value, dict) else {}
            )
            fields = {
                "total_price_and_currency",
                "payment_conditions",
                "terms_and_conditions",
                "shipping_information",
                "return_and_refund_policy",
                "purchasable",
            }
            missing = sorted(field for field in fields if required.get(field) is not True)
            if missing:
                return result(
                    "needs_human_review",
                    f"Landing-page commercial disclosures are unconfirmed: {', '.join(missing)}.",
                    "Confirm all required disclosures and product purchase availability.",
                    "platform_context.landing_page.required_information",
                    missing,
                )
            return result(
                "pass",
                "All supplied landing-page commercial disclosure checks are confirmed.",
                "No action required.",
                "platform_context.landing_page.required_information",
            )

        if rule_id == "GMC-H-015":
            issues = self._editorial_issues(text, include_promotions=False)
            return result(
                "fail" if issues else "pass",
                f"Editorial pattern(s) detected: {', '.join(issues)}."
                if issues
                else "Deterministic capitalization, punctuation, spacing, and repetition checks passed.",
                "Correct the identified editorial patterns."
                if issues
                else "No action required.",
                "content",
                issues,
            )

        if rule_id == "GMC-B-001":
            normalized = _normalize(title)
            has_product = any(
                value in normalized
                for value in ("serum", "sérum", "cleanser", "moisturizer", "cream", "set")
            )
            has_size = bool(re.search(r"\b\d+(?:\.\d+)?\s*m[lL]\b", title))
            if has_product and has_size:
                return result(
                    "pass",
                    "Title includes a product type and size/variant information.",
                    "No action required.",
                    "content.title",
                    title,
                )
            return result(
                "needs_human_review",
                "Recommended title structure is incomplete.",
                "Consider product type, verified feature, and size/variant in the title.",
                "content.title",
                title,
            )

        if rule_id == "INT-LST-001":
            bullets = content.get("bullet_points") or []
            checks = {
                "title_count": 1 if _is_present(title) else 0,
                "bullet_count": len(bullets),
                "description_count": 1 if _is_present(description) else 0,
            }
            passed = checks == {
                "title_count": 1,
                "bullet_count": 5,
                "description_count": 1,
            }
            return result(
                "pass" if passed else "fail",
                "Listing contains one title, five bullets, and one description."
                if passed
                else "Internal listing structure is incomplete.",
                "Provide exactly one title, five bullet points, and one description.",
                "content",
                checks,
            )

        if rule_id == "INT-LST-002":
            missing = [
                claim.get("claim_id")
                for claim in claims
                if not (claim.get("fact_ids") or [])
            ]
            return result(
                "fail" if missing else "pass",
                f"Claims without fact_ids: {missing}."
                if missing
                else "Every declared claim has at least one fact_id.",
                "Add fact_ids or remove unsupported claims."
                if missing
                else "No action required.",
                "claims",
                missing,
            )

        if rule_id.startswith("TTA-") or rule_id.startswith("INT-VID-"):
            return self._evaluate_tiktok_rule(
                rule, payload, fact_report, segments, result
            )

        if rule_id.startswith("INT-SOC-"):
            return self._evaluate_social_rule(rule, payload, segments, result)

        if rule_id == "BRAND-H-001":
            medical_categories = {
                "medical_claim",
                "structure_function",
                "anti_aging",
                "clinical_claim",
            }
            hits = [
                item
                for item in self._matched_lexicon_rules(segments, "prohibited")
                if item.get("category") in medical_categories
            ]
            return result(
                "fail" if hits else "pass",
                f"Medical/structure-function prohibited expression(s) detected: {[item['rule_id'] for item in hits]}."
                if hits
                else "No prohibited medical or structure/function expression was detected.",
                "Remove the medicalized claim and use a supported cosmetic experience statement."
                if hits
                else "No action required.",
                hits[0]["content_location"] if hits else "content",
                [item["matched_text"] for item in hits],
            )

        if rule_id == "BRAND-H-002":
            hits = self._matched_lexicon_rules(segments, "prohibited")
            return result(
                "fail" if hits else "pass",
                f"Prohibited brand expression(s) detected: {[item['rule_id'] for item in hits]}."
                if hits
                else "No miracle, guarantee, perfect, certification, or unsafe absolute expression was detected.",
                "Replace prohibited wording with the approved cautious alternative."
                if hits
                else "No action required.",
                hits[0]["content_location"] if hits else "content",
                [item["matched_text"] for item in hits],
            )

        if rule_id == "INT-MKT-001":
            market_expected: dict[str, str] | None = MARKET_PROFILES.get(
                str(payload.get("market") or "")
            )
            market_currency: Any = _get_path(
                payload, "platform_context.feed_fields.price.currency"
            )
            if market_currency is None:
                market_currency = _get_path(payload, "platform_context.currency")
            issues = []
            if not market_expected:
                issues.append("unsupported_market")
            else:
                if payload.get("language") != market_expected["language"]:
                    issues.append("language_mismatch")
                if market_currency and market_currency != market_expected["currency"]:
                    issues.append("currency_mismatch")
            return result(
                "fail" if issues else "pass",
                f"Market profile issue(s): {', '.join(issues)}."
                if issues
                else "Language and supplied currency match the target-market profile.",
                "Use the configured market language and currency."
                if issues
                else "No action required.",
                "market/language/platform_context",
                issues,
                {"expected": market_expected, "currency": market_currency},
            )

        if rule_id == "INT-REV-001":
            review = payload.get("human_review") or {}
            forbidden = re.search(
                r"\b(platform approved|legally compliant|guaranteed to publish)\b",
                _normalize(text),
            )
            passed = review.get("required") is True and not forbidden
            return result(
                "pass" if passed else "fail",
                "Human final review remains required and no approval guarantee is stated."
                if passed
                else "Human-review status is missing or content claims automatic approval/compliance.",
                "Set human_review.required=true and remove approval or compliance guarantees.",
                "human_review",
                review,
            )

        return result(
            "needs_human_review",
            "No deterministic evaluator is configured for this applicable rule.",
            "Review this rule manually and record supporting evidence.",
            str((rule.get("limit") or {}).get("field") or "content"),
        )

    def _evaluate_tiktok_rule(
        self,
        rule: dict[str, Any],
        payload: dict[str, Any],
        fact_report: dict[str, Any],
        segments: list[dict[str, str]],
        result: Any,
    ) -> dict[str, Any]:
        rule_id = rule["rule_id"]
        content = payload.get("content") or {}
        context = payload.get("platform_context") or {}
        video = context.get("video") or {}
        landing = context.get("landing_page") or {}
        caption = str(content.get("caption") or "")
        duration = (
            video.get("duration_seconds")
            or context.get("duration_seconds")
            or _get_path(payload, "platform_fields.duration_seconds")
        )
        text = "\n".join(item["text"] for item in segments)

        if rule_id == "TTA-H-001":
            hits = self._matched_lexicon_rules(segments, "prohibited")
            fact_blocked = _get_path(fact_report, "export_gate.status") == "blocked"
            failed = bool(hits) or fact_blocked
            return result(
                "fail" if failed else "pass",
                "Blocked fact claims or exaggerated/absolute expressions were detected."
                if failed
                else "No blocked result promise or exaggerated absolute expression was detected.",
                "Resolve fact errors and remove exaggerated or guaranteed result claims."
                if failed
                else "No action required.",
                "content",
                [item["matched_text"] for item in hits],
            )

        if rule_id == "TTA-H-002":
            consistency = landing.get("ad_consistency") or {}
            keys = ["product", "promotion", "price", "discount", "disclaimers", "terms"]
            missing = [key for key in keys if consistency.get(key) is not True]
            return result(
                "needs_human_review" if missing else "pass",
                f"Ad-to-landing consistency is unconfirmed for: {', '.join(missing)}."
                if missing
                else "All supplied ad-to-landing consistency checks are confirmed.",
                "Confirm product and commercial terms against the landing page."
                if missing
                else "No action required.",
                "platform_context.landing_page.ad_consistency",
                missing,
            )

        if rule_id == "TTA-H-003":
            creative = context.get("creative") or {}
            keys = [
                "no_distorted_before_after",
                "no_fake_play_button",
                "no_fake_close_button",
                "no_fake_carousel_indicator",
                "no_fake_cta",
            ]
            missing = [key for key in keys if creative.get(key) is not True]
            return result(
                "needs_human_review" if missing else "pass",
                f"Creative integrity is unconfirmed for: {', '.join(missing)}."
                if missing
                else "All supplied creative-integrity checks are confirmed.",
                "Review the final creative for misleading comparisons or simulated UI."
                if missing
                else "No action required.",
                "platform_context.creative",
                missing,
            )

        if rule_id == "TTA-H-004":
            ai = payload.get("ai_disclosure") or {}
            if not self._is_ai_content(payload):
                return result(
                    "not_applicable",
                    "Content is not marked as AIGC or significantly edited.",
                    "No action required.",
                    "ai_disclosure.aigc_status",
                )
            passed = ai.get("label_required") is True and _is_present(ai.get("method"))
            return result(
                "pass" if passed else "fail",
                "AIGC disclosure requirement and method are recorded."
                if passed
                else "AIGC content lacks a recorded disclosure method.",
                "Set label_required=true and record the platform label or disclosure method.",
                "ai_disclosure",
                ai,
            )

        if rule_id == "TTA-H-005":
            passed = isinstance(duration, (int, float)) and 5 <= duration <= 60
            return result(
                "pass" if passed else "fail",
                f"Video duration is {duration} seconds and within 5–60."
                if passed
                else "Video duration is missing or outside 5–60 seconds.",
                "Set video duration between 5 and 60 seconds.",
                "platform_context.video.duration_seconds",
                duration,
            )

        if rule_id == "TTA-H-006":
            static_share = video.get("static_share_percent")
            checks = {
                "aspect_ratio": video.get("aspect_ratio") in {"9:16", "1:1", "16:9"},
                "audio": video.get("audio_clear") is True,
                "static_share": isinstance(static_share, (int, float))
                and static_share <= 50,
            }
            failed = [key for key, value in checks.items() if not value]
            return result(
                "fail" if failed else "pass",
                f"Video format requirement(s) failed: {', '.join(failed)}."
                if failed
                else "Aspect ratio, audio, and static-share checks passed.",
                "Correct video aspect ratio, audio, or static content share."
                if failed
                else "No action required.",
                "platform_context.video",
                failed,
            )

        if rule_id == "TTA-H-007":
            file_format = str(video.get("file_format") or "").casefold()
            file_size = video.get("file_size_mb")
            bitrate = video.get("bitrate_kbps")
            checks = {
                "format": file_format in {".mp4", ".mov", ".mpeg", ".3gp", ".avi"},
                "size": isinstance(file_size, (int, float)) and file_size <= 500,
                "bitrate": isinstance(bitrate, (int, float)) and bitrate >= 516,
                "duration": isinstance(duration, (int, float)) and duration <= 60,
            }
            failed = [key for key, value in checks.items() if not value]
            return result(
                "fail" if failed else "pass",
                f"Non-Spark file specification(s) failed: {', '.join(failed)}."
                if failed
                else "Non-Spark file format, size, bitrate, and duration checks passed.",
                "Correct the supplied video-file specifications."
                if failed
                else "No action required.",
                "platform_context.video",
                failed,
            )

        if rule_id == "TTA-H-008":
            matches = HASHTAG_OR_AT_PATTERN.findall(caption)
            return result(
                "fail" if matches else "pass",
                f"Caption contains unsupported link, @, or hashtag content: {matches}."
                if matches
                else "Caption contains no clickable link, @ symbol, or hashtag.",
                "Remove links, @ mentions, and hashtags from the Non-Spark caption."
                if matches
                else "No action required.",
                "content.caption",
                matches,
            )

        if rule_id == "TTA-H-009":
            landing_checks: dict[str, Any] = {
                "functional_in_target_market": landing.get(
                    "functional_in_target_market"
                ),
                "complete": landing.get("complete"),
                "mobile_friendly": landing.get("mobile_friendly"),
                "no_automatic_download": landing.get("no_automatic_download"),
                "no_forced_personal_information": landing.get(
                    "no_forced_personal_information"
                ),
            }
            missing = [
                key for key, value in landing_checks.items() if value is not True
            ]
            return result(
                "needs_human_review" if missing else "pass",
                f"Landing-page functionality is unconfirmed for: {', '.join(missing)}."
                if missing
                else "All supplied landing-page functionality checks are confirmed.",
                "Test the landing page in the target market and on mobile."
                if missing
                else "No action required.",
                "platform_context.landing_page",
                missing,
            )

        if rule_id == "TTA-H-010":
            required = landing.get("required_information") or {}
            keys = [
                "contact_details",
                "company_name",
                "company_address",
                "business_license_when_applicable",
                "price_in_local_currency",
                "shipping_information",
                "return_and_refund_policy",
                "terms_and_conditions",
                "privacy_policy",
            ]
            missing = [key for key in keys if required.get(key) is not True]
            return result(
                "needs_human_review" if missing else "pass",
                f"Ecommerce landing-page information is unconfirmed: {', '.join(missing)}."
                if missing
                else "All supplied ecommerce information checks are confirmed.",
                "Confirm all required ecommerce and local-law information."
                if missing
                else "No action required.",
                "platform_context.landing_page.required_information",
                missing,
            )

        if rule_id == "TTA-H-011":
            expected = MARKET_PROFILES.get(str(payload.get("market") or ""), {})
            supplied_currency = context.get("currency")
            issues = []
            if payload.get("language") != expected.get("language"):
                issues.append("language")
            if supplied_currency and supplied_currency != expected.get("currency"):
                issues.append("currency")
            if landing.get("content_context_localized") is not True:
                issues.append("content_context")
            return result(
                "fail" if issues else "pass",
                f"Target-market localization issue(s): {', '.join(issues)}."
                if issues
                else "Language, currency, and context match the target market.",
                "Correct language, currency, or local context."
                if issues
                else "No action required.",
                "market/language/platform_context.currency",
                issues,
            )

        if rule_id == "TTA-H-012":
            issues = self._editorial_issues(text, include_promotions=False)
            return result(
                "fail" if issues else "pass",
                f"Editorial pattern(s) detected: {', '.join(issues)}."
                if issues
                else "Deterministic caption and on-screen editorial checks passed.",
                "Correct capitalization, spacing, symbols, or punctuation."
                if issues
                else "No action required.",
                "content.caption/content.scenes",
                issues,
            )

        if rule_id == "TTA-B-001":
            passed = (
                video.get("aspect_ratio") == "9:16"
                and self._resolution_meets(
                    str(video.get("resolution") or ""), minimum=(540, 960)
                )
            )
            return result(
                "pass" if passed else "needs_human_review",
                "Recommended 9:16 and at least 540×960 are supplied."
                if passed
                else "Recommended vertical aspect ratio or resolution is not confirmed.",
                "Prefer 9:16 and at least 540×960.",
                "platform_context.video",
            )

        if rule_id == "TTA-B-002":
            voiceovers = [
                item["text"]
                for item in segments
                if item["location"].endswith(".voiceover")
            ]
            if not voiceovers:
                return result(
                    "needs_human_review",
                    "No voiceover text is available for naturalness review.",
                    "Provide and review target-market voiceover.",
                    "content.scenes",
                )
            return result(
                "needs_human_review",
                "Voiceover is present, but naturalness requires human language review.",
                "Have a target-market reviewer confirm clarity and naturalness.",
                "content.scenes",
            )

        if rule_id == "INT-VID-001":
            scenes = content.get("scenes") or []
            required = {"timecode", "visual", "voiceover", "on_screen_text", "fact_ids"}
            missing_scene_fields = [
                index
                for index, scene in enumerate(scenes)
                if not isinstance(scene, dict) or not required.issubset(scene)
            ]
            passed = duration in {15, 30} and bool(scenes) and not missing_scene_fields
            return result(
                "pass" if passed else "fail",
                "Duration and scene field structure meet the project template."
                if passed
                else "Project video duration or scene structure is invalid.",
                "Use a 15/30 second duration and include every required scene field.",
                "content.scenes",
                {"duration": duration, "missing_scene_fields": missing_scene_fields},
            )

        if rule_id == "INT-VID-002":
            scenes = content.get("scenes") or []
            roles = [str(scene.get("role") or "") for scene in scenes if isinstance(scene, dict)]
            cta_count = sum(1 for role in roles if role == "cta")
            has_hook = "hook" in roles
            passed = has_hook and cta_count == 1
            return result(
                "pass" if passed else "needs_human_review",
                "Script has a hook and one CTA scene."
                if passed
                else "Recommended Hook-to-CTA sequence is incomplete.",
                "Review the scene sequence and retain one CTA.",
                "content.scenes",
                roles,
            )

        if rule_id == "INT-VID-003":
            missing_scenes: list[int] = []
            for index, scene in enumerate(content.get("scenes") or []):
                if not isinstance(scene, dict):
                    continue
                factual_text = " ".join(
                    str(scene.get(key) or "")
                    for key in ("voiceover", "on_screen_text")
                )
                if self._looks_factual(factual_text) and not (scene.get("fact_ids") or []):
                    missing_scenes.append(index)
            return result(
                "fail" if missing_scenes else "pass",
                f"Factual scene text lacks fact_ids in scene(s): {missing_scenes}."
                if missing_scenes
                else "Factual scene text has fact bindings.",
                "Add fact_ids or remove unsupported scene claims."
                if missing_scenes
                else "No action required.",
                "content.scenes",
                missing_scenes,
            )

        return result(
            "needs_human_review",
            "TikTok rule requires manual review.",
            "Review the final ad and landing page.",
            "content",
        )

    def _evaluate_social_rule(
        self,
        rule: dict[str, Any],
        payload: dict[str, Any],
        segments: list[dict[str, str]],
        result: Any,
    ) -> dict[str, Any]:
        rule_id = rule["rule_id"]
        content = payload.get("content") or {}
        context = payload.get("platform_context") or {}
        if rule_id == "INT-SOC-001":
            scope_note = str(context.get("platform_scope_note") or "")
            approval_claim = re.search(
                r"\b(?:instagram|facebook|x|pinterest)\s+(?:approved|compliant)\b",
                _normalize("\n".join(item["text"] for item in segments)),
            )
            passed = payload.get("platform") == "generic_social" and not approval_claim
            return result(
                "pass" if passed else "fail",
                "Generic Social scope is retained without a platform-approval claim."
                if passed
                else "Generic Social output claims or implies real-platform approval.",
                "Keep generic_social scope and run real platform rules before publishing.",
                "platform/platform_context.platform_scope_note",
                scope_note,
            )

        if rule_id == "INT-SOC-002":
            claims = payload.get("claims") or []
            verified = [claim for claim in claims if claim.get("fact_ids")]
            checks = {
                "hook": _is_present(content.get("hook")),
                "body": _is_present(content.get("body")),
                "cta": _is_present(content.get("cta")),
                "ai_disclosure": isinstance(payload.get("ai_disclosure"), dict),
                "verified_benefits": 1 <= len(verified) <= 2,
            }
            failed = [key for key, value in checks.items() if not value]
            return result(
                "fail" if failed else "pass",
                f"Social content section check failed: {', '.join(failed)}."
                if failed
                else "Required social sections and verified benefit count are present.",
                "Provide Hook, body, one CTA, AI field, and 1–2 fact-bound benefits."
                if failed
                else "No action required.",
                "content/claims/ai_disclosure",
                failed,
            )

        if rule_id == "INT-SOC-003":
            commercial = context.get("conditional_commercial_claims") or {}
            unsupported = [
                key
                for key, value in commercial.items()
                if _is_present(value)
                and not (
                    isinstance(value, dict)
                    and value.get("fact_supported") is True
                )
            ]
            return result(
                "fail" if unsupported else "pass",
                f"Commercial claims lack fact support: {', '.join(unsupported)}."
                if unsupported
                else "No unsupported conditional commercial claim is supplied.",
                "Remove the claim or attach current fact support."
                if unsupported
                else "No action required.",
                "platform_context.conditional_commercial_claims",
                unsupported,
            )

        return result(
            "needs_human_review",
            "Generic Social rule requires manual review.",
            "Select a real publication platform and rerun current rules.",
            "platform_context",
        )

    def _check_lexicon(
        self, payload: dict[str, Any], segments: list[dict[str, str]]
    ) -> list[dict[str, Any]]:
        hits = self._matched_lexicon_rules(segments)
        if not hits:
            return [
                self._result(
                    rule_id="BRAND-LEXICON",
                    rule_type="brand_rule",
                    severity="block",
                    status="pass",
                    content_location="content",
                    matched_text=None,
                    reason="No prohibited or caution lexicon item was detected.",
                    source_url="data/brand/prohibited_terms.csv",
                    suggested_action="No action required.",
                )
            ]
        output = []
        for hit in hits:
            severity = hit.get("severity") or "review"
            output.append(
                self._result(
                    rule_id=hit["rule_id"],
                    rule_type="brand_rule",
                    severity=severity,
                    status="fail" if severity == "block" else "needs_human_review",
                    content_location=hit["content_location"],
                    matched_text=hit["matched_text"],
                    reason=hit.get("reason_zh") or "Brand lexicon rule matched.",
                    source_url=hit.get("source_url"),
                    suggested_action=(
                        hit.get(
                            "safe_alternative_es"
                            if _language_family(str(payload.get("language") or ""))
                            == "es"
                            else "safe_alternative_en"
                        )
                        or "Remove or revise the expression."
                    ),
                    evidence={
                        "category": hit.get("category"),
                        "evidence_required": hit.get("evidence_required"),
                    },
                )
            )
        return output

    def _matched_lexicon_rules(
        self,
        segments: list[dict[str, str]],
        list_type: str | None = None,
    ) -> list[dict[str, Any]]:
        hits: list[dict[str, Any]] = []
        seen: set[tuple[str, str]] = set()
        for rule in self.lexicon_rules:
            if list_type and rule.get("list_type") != list_type:
                continue
            variants = _split_term_variants(rule.get("term_en")) + _split_term_variants(
                rule.get("term_es")
            )
            for segment in segments:
                normalized = _normalize(segment["text"])
                for variant in variants:
                    if _term_pattern(variant).search(normalized):
                        key = (rule["rule_id"], segment["location"])
                        if key not in seen:
                            hits.append(
                                {
                                    **rule,
                                    "content_location": segment["location"],
                                    "matched_text": variant,
                                }
                            )
                            seen.add(key)
                        break
        return hits

    def _check_terminology(
        self, payload: dict[str, Any], segments: list[dict[str, str]]
    ) -> dict[str, Any]:
        language = _language_family(str(payload.get("language") or ""))
        accepted_variants: set[str] = set()
        for term in self.terminology:
            accepted_variants.add(
                _normalize(
                    term.get("preferred_es" if language == "es" else "preferred_en")
                )
            )
            accepted_variants.update(
                _normalize(value)
                for value in _split_term_variants(term.get("allowed_variants"))
            )
        accepted_variants.discard("")
        violations: list[dict[str, str]] = []
        for term in self.terminology:
            variants = _split_term_variants(term.get("avoid_terms"))
            for variant in variants:
                if _normalize(variant) in accepted_variants:
                    continue
                if (
                    term.get("term_id") == "TERM-002"
                    and _normalize(variant) == "serum"
                    and language == "en"
                ):
                    continue
                for segment in segments:
                    normalized = _normalize(segment["text"])
                    if _term_pattern(variant).search(normalized):
                        violations.append(
                            {
                                "term_id": str(term.get("term_id")),
                                "location": segment["location"],
                                "matched_text": variant,
                                "preferred": str(
                                    term.get(
                                        "preferred_es"
                                        if language == "es"
                                        else "preferred_en"
                                    )
                                    or ""
                                ),
                                "reason": str(term.get("reason") or ""),
                            }
                        )
        if violations:
            first = violations[0]
            return self._result(
                rule_id="TERM-CONSISTENCY",
                rule_type="terminology_rule",
                severity="block",
                status="fail",
                content_location=first["location"],
                matched_text=[item["matched_text"] for item in violations],
                reason=f"Non-preferred or avoided terminology detected in {len(violations)} location(s).",
                source_url="data/brand/terminology.xlsx",
                suggested_action=f"Use the preferred term: {first['preferred']}.",
                evidence={"violations": violations},
            )
        return self._result(
            rule_id="TERM-CONSISTENCY",
            rule_type="terminology_rule",
            severity="block",
            status="pass",
            content_location="content",
            matched_text=None,
            reason="No avoided terminology variant was detected for the target language.",
            source_url="data/brand/terminology.xlsx",
            suggested_action="No action required.",
        )

    def _check_cta(self, payload: dict[str, Any]) -> dict[str, Any]:
        content_type = str(payload.get("content_type") or "")
        if content_type == "product_listing":
            return self._result(
                rule_id="BRAND-CTA-001",
                rule_type="brand_rule",
                severity="review",
                status="not_applicable",
                content_location="content.cta",
                matched_text=None,
                reason="A standalone CTA is not required for the product-listing content block.",
                source_url="data/brand/brand_voice_guide.md",
                suggested_action="No action required.",
            )
        cta = str(_get_path(payload, "content.cta") or "")
        if not cta:
            scenes = _get_path(payload, "content.scenes", []) or []
            cta_values = [
                str(scene.get("voiceover") or scene.get("on_screen_text") or "")
                for scene in scenes
                if isinstance(scene, dict) and scene.get("role") == "cta"
            ]
            cta = " ".join(value for value in cta_values if value)
        normalized = _normalize(cta).rstrip(".!?")
        pressure = [
            pattern
            for pattern in HIGH_PRESSURE_CTA_PATTERNS
            if re.search(pattern, normalized, re.IGNORECASE)
        ]
        if pressure:
            return self._result(
                rule_id="BRAND-CTA-001",
                rule_type="brand_rule",
                severity="block",
                status="fail",
                content_location="content.cta",
                matched_text=cta,
                reason="CTA uses high-pressure, result-guarantee, scarcity, or endorsement language.",
                source_url="data/brand/brand_voice_guide.md",
                suggested_action="Use an invitation CTA such as 'See product details' or 'Consulta los detalles'.",
            )
        if not cta:
            return self._result(
                rule_id="BRAND-CTA-001",
                rule_type="brand_rule",
                severity="review",
                status="needs_human_review",
                content_location="content.cta",
                matched_text=None,
                reason="A CTA is required for this content type but none was found.",
                source_url="data/brand/brand_voice_guide.md",
                suggested_action="Add one invitation-style CTA.",
            )
        language = _language_family(str(payload.get("language") or ""))
        recommended = normalized in SAFE_CTAS[language]
        return self._result(
            rule_id="BRAND-CTA-001",
            rule_type="brand_rule",
            severity="review",
            status="pass" if recommended else "needs_human_review",
            content_location="content.cta",
            matched_text=cta,
            reason="CTA matches the approved brand list."
            if recommended
            else "CTA is not prohibited, but it is outside the approved phrase list.",
            source_url="data/brand/brand_voice_guide.md",
            suggested_action="No action required."
            if recommended
            else "Confirm tone manually or use an approved invitation CTA.",
        )

    def _check_ai_disclosure(self, payload: dict[str, Any]) -> dict[str, Any]:
        ai = payload.get("ai_disclosure")
        if not isinstance(ai, dict):
            return self._result(
                rule_id="INT-AI-001",
                rule_type="project_internal_rule",
                severity="block",
                status="fail",
                content_location="ai_disclosure",
                matched_text=None,
                reason="Required AI-origin and disclosure fields are missing.",
                source_url="data/platform_rules/platform_rules.json",
                suggested_action="Add ai_disclosure with origin, requirement, and method fields.",
            )
        required_keys = {"aigc_status", "label_required", "method"}
        missing = sorted(key for key in required_keys if key not in ai)
        if missing:
            return self._result(
                rule_id="INT-AI-001",
                rule_type="project_internal_rule",
                severity="block",
                status="fail",
                content_location="ai_disclosure",
                matched_text=missing,
                reason=f"AI disclosure field(s) are missing: {', '.join(missing)}.",
                source_url="data/platform_rules/platform_rules.json",
                suggested_action="Populate every required AI disclosure field.",
            )
        return self._result(
            rule_id="INT-AI-001",
            rule_type="project_internal_rule",
            severity="block",
            status="pass",
            content_location="ai_disclosure",
            matched_text=ai.get("aigc_status"),
            reason="AI origin, label requirement, and method fields are present.",
            source_url="data/platform_rules/platform_rules.json",
            suggested_action="No action required.",
        )

    def _check_brand_tone(
        self, payload: dict[str, Any], segments: list[dict[str, str]]
    ) -> dict[str, Any]:
        issues = []
        text = "\n".join(segment["text"] for segment in segments)
        exclamations = text.count("!") + text.count("！")
        if exclamations > 1:
            issues.append("more_than_one_exclamation_mark")
        if ALL_CAPS_PATTERN.search(text):
            issues.append("all_caps")
        if any(len(re.split(r"(?<=[.!?])\s+", segment["text"])) == 1 and len(segment["text"]) > 300 for segment in segments):
            issues.append("overlong_sentence_or_block")
        return self._result(
            rule_id="BRAND-TONE-001",
            rule_type="brand_rule",
            severity="review",
            status="needs_human_review" if issues else "pass",
            content_location="content",
            matched_text=issues,
            reason=f"Brand-tone heuristic issue(s): {', '.join(issues)}."
            if issues
            else "Deterministic calmness and concision indicators passed.",
            source_url="data/brand/brand_voice_guide.md",
            suggested_action="Reduce excitement, capitalization, or sentence density."
            if issues
            else "No action required; retain human tone review.",
        )

    @staticmethod
    def _result(
        *,
        rule_id: str,
        rule_type: str,
        severity: str,
        status: str,
        content_location: str,
        matched_text: Any,
        reason: str,
        source_url: str | None,
        suggested_action: str,
        evidence: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        if status not in ALLOWED_STATUSES:
            raise ValueError(f"Unsupported rule status: {status}")
        return {
            "rule_id": rule_id,
            "rule_type": rule_type,
            "severity": severity,
            "status": status,
            "content_location": content_location,
            "matched_text": matched_text,
            "reason": reason,
            "source_url": source_url,
            "suggested_action": suggested_action,
            "evidence": evidence or {},
        }

    @staticmethod
    def _is_ai_text(payload: dict[str, Any]) -> bool:
        value = _normalize(_get_path(payload, "ai_disclosure.aigc_status"))
        return "generative" in value or "aigc" in value or value == "ai"

    @staticmethod
    def _is_ai_image(payload: dict[str, Any]) -> bool:
        value = _normalize(
            _get_path(payload, "platform_context.image.ai_origin")
            or _get_path(payload, "platform_context.image_ai_origin")
        )
        return value in {"generative_ai", "aigc", "significantly_edited", "ai"}

    def _is_ai_content(self, payload: dict[str, Any]) -> bool:
        return self._is_ai_text(payload) or self._is_ai_image(payload)

    @staticmethod
    def _promotion_matches(text: str) -> list[str]:
        return [
            pattern
            for pattern in PROMOTIONAL_PATTERNS
            if re.search(pattern, text, re.IGNORECASE)
        ]

    def _editorial_issues(
        self, text: str, *, include_promotions: bool
    ) -> list[str]:
        issues: list[str] = []
        if ALL_CAPS_PATTERN.search(text):
            issues.append("excessive_capitalization")
        if EXCESSIVE_PUNCTUATION_PATTERN.search(text):
            issues.append("excessive_punctuation")
        if GIMMICKY_SYMBOL_PATTERN.search(text):
            issues.append("gimmicky_symbols")
        if ABNORMAL_SPACING_PATTERN.search(text):
            issues.append("abnormal_spacing")
        if self._keyword_stuffing(text):
            issues.append("keyword_repetition")
        if include_promotions:
            issues.extend(self._promotion_matches(text))
        return list(dict.fromkeys(issues))

    @staticmethod
    def _keyword_stuffing(text: str) -> bool:
        words = [
            word
            for word in re.findall(r"[A-Za-zÀ-ÿ]{3,}", _normalize(text))
            if word not in STOPWORDS
        ]
        if len(words) < 8:
            return False
        counts = Counter(words)
        return any(count >= 4 and count / len(words) >= 0.25 for count in counts.values())

    @staticmethod
    def _resolution_meets(value: str, minimum: tuple[int, int]) -> bool:
        match = re.fullmatch(r"\s*(\d+)\s*[x×]\s*(\d+)\s*", value)
        if not match:
            return False
        width, height = (int(match.group(1)), int(match.group(2)))
        return width >= minimum[0] and height >= minimum[1]

    @staticmethod
    def _looks_factual(text: str) -> bool:
        normalized = _normalize(text)
        cues = (
            "serum",
            "sérum",
            "contains",
            "made with",
            "fragrance",
            "apply",
            "pump",
            "ml",
            "helps",
            "contiene",
            "sin fragancia",
            "aplica",
            "dosis",
            "ayuda",
        )
        return bool(re.search(r"\d", normalized)) or any(cue in normalized for cue in cues)

    @staticmethod
    def _summarize(results: list[dict[str, Any]]) -> dict[str, Any]:
        by_type: dict[str, Counter[str]] = defaultdict(Counter)
        for item in results:
            by_type[item["rule_type"]][item["status"]] += 1
        summary_by_type = {}
        for rule_type, counts in sorted(by_type.items()):
            summary_by_type[rule_type] = {
                status: counts.get(status, 0)
                for status in ("pass", "fail", "not_applicable", "needs_human_review")
            }
            summary_by_type[rule_type]["total"] = sum(counts.values())
        block_failures = sum(
            1
            for item in results
            if item["severity"] == "block" and item["status"] == "fail"
        )
        unresolved_hard = sum(
            1
            for item in results
            if item["rule_type"] == "platform_hard_rule"
            and item["status"] == "needs_human_review"
        )
        return {
            "by_rule_type": summary_by_type,
            "totals": {
                "evaluated": len(results),
                "pass": sum(item["status"] == "pass" for item in results),
                "fail": sum(item["status"] == "fail" for item in results),
                "not_applicable": sum(
                    item["status"] == "not_applicable" for item in results
                ),
                "needs_human_review": sum(
                    item["status"] == "needs_human_review" for item in results
                ),
                "block_failures": block_failures,
                "unresolved_hard_rules": unresolved_hard,
            },
        }

    def _score_quality(
        self,
        payload: dict[str, Any],
        fact_report: dict[str, Any],
        results: list[dict[str, Any]],
        segments: list[dict[str, str]],
    ) -> dict[str, Any]:
        dimensions: dict[str, dict[str, Any]] = {}

        def dimension(
            name: str,
            deductions: list[dict[str, Any]],
            explanation: str,
            mode: str,
        ) -> None:
            total_deduction = min(100, sum(item["points"] for item in deductions))
            score = max(0, 100 - total_deduction)
            weight = QUALITY_WEIGHTS[name]
            dimensions[name] = {
                "weight": weight,
                "hard_gate": name == "fact_accuracy",
                "assessment_mode": mode,
                "score": score,
                "weighted_points": round(score * weight / 100, 2),
                "deductions": deductions,
                "explanation": explanation,
            }

        fact_summary = fact_report.get("summary") or {}
        error_rate = float(fact_summary.get("fact_error_rate") or 0)
        partial = int(fact_summary.get("partially_supported_count") or 0)
        fact_deductions = []
        if error_rate:
            fact_deductions.append(
                {
                    "points": round(min(100, error_rate * 100), 2),
                    "reason": "Unsupported or contradicted factual claims reduce fact accuracy.",
                    "evidence": {"fact_error_rate": error_rate},
                }
            )
        if partial:
            fact_deductions.append(
                {
                    "points": min(20, partial * 5),
                    "reason": "Partially supported claims require qualification or evidence.",
                    "evidence": {"partially_supported_count": partial},
                }
            )
        dimension(
            "fact_accuracy",
            fact_deductions,
            "Derived only from the upstream fact-check snapshot; the rule checker does not rejudge facts.",
            "deterministic_from_fact_gate",
        )

        platform_items = [
            item
            for item in results
            if item["rule_type"]
            in {"platform_hard_rule", "platform_best_practice", "project_internal_rule"}
        ]
        platform_deductions = []
        for item in platform_items:
            points = 0
            if item["status"] == "fail":
                points = 15 if item["severity"] == "block" else 8
            elif item["status"] == "needs_human_review":
                points = 4
            if points:
                platform_deductions.append(
                    {
                        "points": points,
                        "reason": item["reason"],
                        "rule_id": item["rule_id"],
                    }
                )
        dimension(
            "platform_fit",
            platform_deductions,
            "Based on platform and project-structure results; unresolved checks remain visible.",
            "deterministic_rules",
        )

        editorial = self._editorial_issues(
            "\n".join(item["text"] for item in segments),
            include_promotions=False,
        )
        language_deductions = [
            {
                "points": 8,
                "reason": f"Editorial indicator detected: {issue}.",
                "evidence": issue,
            }
            for issue in editorial
        ]
        language = _language_family(str(payload.get("language") or ""))
        words = set(
            re.findall(
                r"[A-Za-zÀ-ÿ]+",
                _normalize("\n".join(item["text"] for item in segments)),
            )
        )
        if language == "es" and len(words & ENGLISH_WORDS) > len(words & SPANISH_WORDS):
            language_deductions.append(
                {
                    "points": 20,
                    "reason": "English lexical indicators dominate an es-MX output.",
                    "evidence": sorted(words & ENGLISH_WORDS),
                }
            )
        if language == "en" and len(words & SPANISH_WORDS) > len(words & ENGLISH_WORDS):
            language_deductions.append(
                {
                    "points": 20,
                    "reason": "Spanish lexical indicators dominate an en-US output.",
                    "evidence": sorted(words & SPANISH_WORDS),
                }
            )
        dimension(
            "language_naturalness",
            language_deductions,
            "Heuristic language and editorial precheck; native-speaker review is still required.",
            "heuristic_precheck",
        )

        brand_deductions = []
        for item in results:
            if item["rule_type"] not in {"brand_rule", "terminology_rule"}:
                continue
            points = 0
            if item["status"] == "fail":
                points = 20 if item["severity"] == "block" else 10
            elif item["status"] == "needs_human_review":
                points = 6
            if points:
                brand_deductions.append(
                    {
                        "points": points,
                        "reason": item["reason"],
                        "rule_id": item["rule_id"],
                    }
                )
        dimension(
            "brand_consistency",
            brand_deductions,
            "Brand, terminology, tone, and CTA results are scored separately from platform hard rules.",
            "deterministic_and_heuristic",
        )

        localization_deductions = []
        expected = MARKET_PROFILES.get(str(payload.get("market") or ""))
        if not expected:
            localization_deductions.append(
                {
                    "points": 40,
                    "reason": "Target market is unsupported.",
                    "evidence": payload.get("market"),
                }
            )
        elif payload.get("language") != expected["language"]:
            localization_deductions.append(
                {
                    "points": 35,
                    "reason": "Language does not match the target market.",
                    "evidence": {
                        "actual": payload.get("language"),
                        "expected": expected["language"],
                    },
                }
            )
        full_text = _normalize("\n".join(item["text"] for item in segments))
        if language == "es" and re.search(r"\bvosotros\b|\bvuestro(?:s|as)?\b", full_text):
            localization_deductions.append(
                {
                    "points": 15,
                    "reason": "Spain-oriented vosotros forms are not appropriate for es-MX.",
                    "evidence": "vosotros/vuestro",
                }
            )
        currency = _get_path(payload, "platform_context.feed_fields.price.currency")
        if expected and currency and currency != expected["currency"]:
            localization_deductions.append(
                {
                    "points": 25,
                    "reason": "Currency does not match the target market.",
                    "evidence": {"actual": currency, "expected": expected["currency"]},
                }
            )
        dimension(
            "localization_quality",
            localization_deductions,
            "Checks market-language profile, currency, and selected regional-language indicators.",
            "deterministic_and_heuristic",
        )

        content = payload.get("content") or {}
        marketing_deductions = []
        if not _is_present(content.get("title") or content.get("hook")):
            marketing_deductions.append(
                {
                    "points": 25,
                    "reason": "No clear title or hook is present.",
                    "evidence": "content.title/content.hook",
                }
            )
        benefit_claims = [
            claim
            for claim in payload.get("claims") or []
            if claim.get("evidence_level") == "B"
        ]
        if not benefit_claims:
            marketing_deductions.append(
                {
                    "points": 15,
                    "reason": "No cautious, fact-bound benefit claim is declared.",
                    "evidence": "claims[].evidence_level",
                }
            )
        if payload.get("content_type") in {"short_video_script", "social_ad_copy"}:
            cta_result = next(
                (item for item in results if item["rule_id"] == "BRAND-CTA-001"),
                None,
            )
            if cta_result and cta_result["status"] != "pass":
                marketing_deductions.append(
                    {
                        "points": 15,
                        "reason": cta_result["reason"],
                        "rule_id": cta_result["rule_id"],
                    }
                )
        dimension(
            "marketing_persuasiveness",
            marketing_deductions,
            "Checks presence of an opening, cautious benefit, and content-appropriate CTA.",
            "heuristic_precheck",
        )

        weighted_score = round(
            sum(item["weighted_points"] for item in dimensions.values()), 2
        )
        return {
            "rubric_version": "LF-QUALITY-1.0",
            "weights_total": sum(QUALITY_WEIGHTS.values()),
            "dimensions": dimensions,
            "weighted_score": weighted_score,
            "hard_gate_independent": True,
            "quality_score_can_override_hard_gate": False,
            "interpretation": (
                "Auxiliary precheck score only; not a compliance or publication score."
            ),
        }

    @staticmethod
    def _export_gate(
        fact_report: dict[str, Any],
        results: list[dict[str, Any]],
        quality: dict[str, Any],
    ) -> dict[str, Any]:
        fact_status = _get_path(fact_report, "export_gate.status")
        fact_errors = int(
            (_get_path(fact_report, "summary.unsupported_count") or 0)
            + (_get_path(fact_report, "summary.contradicted_count") or 0)
        )
        block_failures = [
            item
            for item in results
            if item["severity"] == "block" and item["status"] == "fail"
        ]
        review_items = [
            item
            for item in results
            if item["status"] == "needs_human_review"
        ]
        if fact_status == "blocked" or fact_errors:
            status = "blocked"
            priority = "fact_error"
            reason = (
                "Upstream fact errors have the highest blocking priority and must "
                "be resolved before any score or rule result can allow export."
            )
        elif block_failures:
            status = "blocked"
            priority = "hard_rule_failure"
            reason = f"{len(block_failures)} block-level rule failure(s) prevent export."
        elif review_items:
            status = "needs_human_review"
            priority = "unresolved_review"
            reason = f"{len(review_items)} rule item(s) require human confirmation."
        else:
            status = "pass"
            priority = "human_final_review"
            reason = (
                "Automated fact and rule prechecks passed; export remains disabled "
                "until human final review."
            )
        return {
            "status": status,
            "export_allowed": False,
            "blocking_priority": priority,
            "reason": reason,
            "fact_error_has_highest_priority": True,
            "quality_score": quality["weighted_score"],
            "quality_score_can_override": False,
            "human_final_review_required": True,
        }


def check_content_file(
    content_path: str | Path,
    fact_report_path: str | Path,
    platform_rules_path: str | Path,
    terminology_path: str | Path,
    prohibited_terms_path: str | Path,
    previous_content_path: str | Path | None = None,
) -> dict[str, Any]:
    checker = RuleChecker(
        platform_rules_path=platform_rules_path,
        terminology_path=terminology_path,
        prohibited_terms_path=prohibited_terms_path,
    )
    previous = _load_json(previous_content_path) if previous_content_path else None
    return checker.check(
        _load_json(content_path),
        _load_json(fact_report_path),
        previous_content_output=previous,
    )


def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Run LocalizeFlow brand, terminology, and platform rule checks."
    )
    parser.add_argument("--content", required=True)
    parser.add_argument("--fact-report", required=True)
    parser.add_argument(
        "--rules", default="data/platform_rules/platform_rules.json"
    )
    parser.add_argument(
        "--terminology", default="data/brand/terminology.xlsx"
    )
    parser.add_argument(
        "--prohibited-terms", default="data/brand/prohibited_terms.csv"
    )
    parser.add_argument("--previous-content")
    parser.add_argument("--output", required=True)
    args = parser.parse_args(list(argv) if argv is not None else None)

    report = check_content_file(
        content_path=args.content,
        fact_report_path=args.fact_report,
        platform_rules_path=args.rules,
        terminology_path=args.terminology,
        prohibited_terms_path=args.prohibited_terms,
        previous_content_path=args.previous_content,
    )
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return 2 if report["export_gate"]["status"] == "blocked" else 0


if __name__ == "__main__":
    raise SystemExit(main())
