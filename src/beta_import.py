"""Secure, non-persistent import and confirmation gates for Closed Beta SKU facts."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
from collections import defaultdict
from datetime import UTC, datetime
from typing import Any, Iterable

from openpyxl import load_workbook

MAX_UPLOAD_BYTES = 2 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 10 * 1024 * 1024
MAX_ZIP_ENTRIES = 100
MAX_ZIP_RATIO = 100
MAX_ROWS = 50
MAX_COLUMNS = 11
XLSX_SHEET = "SKU Facts"

COLUMNS = (
    "sku",
    "attribute",
    "value",
    "unit",
    "evidence_level",
    "source",
    "source_type",
    "market_scope",
    "allowed_expression",
    "prohibited_expression",
    "generation_policy",
)
ATTRIBUTES = {
    "product_name",
    "specification",
    "ingredient",
    "usage_instruction",
    "packaging_container",
    "packaging_material",
    "packaging_capacity",
    "allowed_claim",
    "prohibited_claim",
}
REQUIRED_ATTRIBUTES = ATTRIBUTES
EVIDENCE_LEVELS = {"A", "B", "C", "U"}
SOURCE_TYPES = {
    "primary_spec",
    "label",
    "testing_report",
    "brand_policy",
    "legal_review",
    "participant_statement",
    "unknown",
}
GENERATION_POLICIES = {"direct", "cautious", "blocked", "not_directly_usable"}
SINGLE_VALUE_ATTRIBUTES = {
    "product_name",
    "specification",
    "usage_instruction",
    "packaging_container",
    "packaging_material",
    "packaging_capacity",
}
SKU_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,63}$")
PROJECT_PATTERN = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,63}$")
DANGEROUS_PREFIXES = ("=", "+", "-", "@")


class BetaImportError(ValueError):
    """Raised when an upload cannot be parsed safely."""


def _issue(
    severity: str,
    code: str,
    message: str,
    *,
    row: int | None = None,
    field: str | None = None,
) -> dict[str, Any]:
    return {"severity": severity, "code": code, "message": message, "row": row, "field": field}


def _safe_text(value: Any, *, max_length: int, field: str, row: int) -> tuple[str, list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    if value is None:
        return "", issues
    if isinstance(value, (dict, list, tuple, set)):
        return "", [_issue("error", "non_scalar", f"{field} 必须是单个值。", row=row, field=field)]
    text = str(value).strip()
    if len(text) > max_length:
        issues.append(_issue("error", "value_too_long", f"{field} 超过 {max_length} 字符。", row=row, field=field))
    if any(ord(character) < 32 and character not in "\t\n\r" for character in text):
        issues.append(_issue("error", "control_character", f"{field} 包含控制字符。", row=row, field=field))
    if text.startswith(DANGEROUS_PREFIXES):
        issues.append(_issue("error", "formula_prefix", f"{field} 以电子表格公式危险字符开头。", row=row, field=field))
    return text, issues


def _inspect_xlsx_archive(data: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ZIP_ENTRIES:
                raise BetaImportError("XLSX 压缩包条目过多。")
            total = 0
            for entry in entries:
                normalized = entry.filename.replace("\\", "/")
                lowered = normalized.lower()
                if normalized.startswith("/") or ".." in normalized.split("/"):
                    raise BetaImportError("XLSX 包含不安全路径。")
                if any(token in lowered for token in ("vbaproject", "externallinks/", "embeddings/", "oleobjects/")) or lowered.endswith(".bin"):
                    raise BetaImportError("XLSX 包含宏、外部链接或嵌入对象。")
                total += entry.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise BetaImportError("XLSX 解压后超过大小限制。")
                if entry.compress_size == 0 and entry.file_size > 0:
                    raise BetaImportError("XLSX 条目压缩信息异常。")
                if entry.compress_size and entry.file_size / entry.compress_size > MAX_ZIP_RATIO:
                    raise BetaImportError("XLSX 压缩比异常。")
    except zipfile.BadZipFile as error:
        raise BetaImportError("XLSX 不是有效的 Office 压缩包。") from error


def _rows_from_xlsx(data: bytes) -> list[dict[str, Any]]:
    _inspect_xlsx_archive(data)
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False, keep_links=False)
    except Exception as error:
        raise BetaImportError("XLSX 无法安全读取。") from error
    try:
        if XLSX_SHEET not in workbook.sheetnames:
            raise BetaImportError(f"XLSX 缺少工作表：{XLSX_SHEET}。")
        sheet = workbook[XLSX_SHEET]
        if (sheet.max_column or 0) > MAX_COLUMNS:
            raise BetaImportError("XLSX 列数超过限制。")
        header = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=4, max_row=4))]
        if tuple(header) != COLUMNS:
            raise BetaImportError("XLSX 列名或顺序与模板不一致。")
        rows: list[dict[str, Any]] = []
        for cells in sheet.iter_rows(min_row=5, max_row=MAX_ROWS + 4, max_col=MAX_COLUMNS):
            if any(cell.data_type == "f" for cell in cells):
                raise BetaImportError(f"XLSX 第 {cells[0].row} 行包含公式。")
            values = [cell.value for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(dict(zip(COLUMNS, values, strict=True)))
        return rows
    finally:
        workbook.close()


def _rows_from_csv(data: bytes) -> list[dict[str, Any]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise BetaImportError("CSV 必须使用 UTF-8 编码。") from error
    try:
        reader = csv.DictReader(io.StringIO(text), strict=True)
        if tuple(reader.fieldnames or ()) != COLUMNS:
            raise BetaImportError("CSV 列名或顺序与模板不一致。")
        rows = list(reader)
    except (csv.Error, UnicodeError) as error:
        raise BetaImportError("CSV 格式无效。") from error
    if len(rows) > MAX_ROWS:
        raise BetaImportError("CSV 行数超过限制。")
    return rows


def _rows_from_json(data: bytes) -> list[dict[str, Any]]:
    try:
        payload = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise BetaImportError("JSON 必须是有效的 UTF-8 JSON。") from error
    if not isinstance(payload, dict) or payload.get("schema_version") != "1.0.0":
        raise BetaImportError("JSON schema_version 必须为 1.0.0。")
    if set(payload) != {"schema_version", "facts"} or not isinstance(payload["facts"], list):
        raise BetaImportError("JSON 只能包含 schema_version 和 facts 数组。")
    if len(payload["facts"]) > MAX_ROWS:
        raise BetaImportError("JSON facts 数量超过限制。")
    if not all(isinstance(row, dict) for row in payload["facts"]):
        raise BetaImportError("JSON facts 中每一项必须是对象。")
    return payload["facts"]


def _normalize_rows(rows: Iterable[dict[str, Any]], project_id: str) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    facts: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for index, raw in enumerate(rows, start=5):
        extra = set(raw) - set(COLUMNS)
        if extra:
            issues.append(_issue("error", "unknown_columns", f"存在未知字段：{', '.join(sorted(extra))}。", row=index))
        normalized: dict[str, str] = {}
        limits = {"value": 2000, "source": 500, "allowed_expression": 1000, "prohibited_expression": 1000}
        for column in COLUMNS:
            text, text_issues = _safe_text(raw.get(column), max_length=limits.get(column, 64), field=column, row=index)
            normalized[column] = text
            issues.extend(text_issues)

        sku = normalized["sku"]
        attribute = normalized["attribute"]
        value = normalized["value"]
        if not SKU_PATTERN.fullmatch(sku):
            issues.append(_issue("error", "invalid_sku", "SKU 格式无效。", row=index, field="sku"))
        if attribute not in ATTRIBUTES:
            issues.append(_issue("error", "invalid_attribute", "attribute 不在允许列表中。", row=index, field="attribute"))
        if not value:
            issues.append(_issue("error", "missing_value", "value 不能为空。", row=index, field="value"))
        if normalized["evidence_level"] not in EVIDENCE_LEVELS:
            issues.append(_issue("error", "invalid_evidence", "evidence_level 必须为 A/B/C/U。", row=index, field="evidence_level"))
        if len(normalized["source"]) < 2:
            issues.append(_issue("error", "missing_source", "每条事实必须有来源。", row=index, field="source"))
        if normalized["source_type"] not in SOURCE_TYPES:
            issues.append(_issue("error", "invalid_source_type", "source_type 无效。", row=index, field="source_type"))
        markets = normalized["market_scope"].split(";") if normalized["market_scope"] else []
        if not markets or set(markets) - {"US", "MX"}:
            issues.append(_issue("error", "invalid_market_scope", "market_scope 必须为 US、MX 或 US;MX。", row=index, field="market_scope"))
        if normalized["generation_policy"] not in GENERATION_POLICIES:
            issues.append(_issue("error", "invalid_policy", "generation_policy 无效。", row=index, field="generation_policy"))
        unknown = value.lower() == "unknown" or normalized["evidence_level"] == "U"
        if unknown and normalized["generation_policy"] not in {"blocked", "not_directly_usable"}:
            issues.append(_issue("error", "unknown_generation", "unknown 字段不得设置为可生成。", row=index, field="generation_policy"))
        if unknown:
            issues.append(_issue("warning", "unknown_value", "字段为 unknown；系统会省略该事实，不会推断。", row=index, field="value"))
        elif normalized["evidence_level"] == "C" or normalized["source_type"] in {"participant_statement", "unknown"}:
            issues.append(_issue("warning", "low_evidence", "低证据字段需要人工复核。", row=index, field="evidence_level"))
        if attribute == "packaging_capacity" and not unknown and not normalized["unit"]:
            issues.append(_issue("warning", "missing_unit", "包装容量缺少单位。", row=index, field="unit"))
        if attribute == "prohibited_claim" and normalized["generation_policy"] != "blocked":
            issues.append(_issue("error", "prohibited_policy", "禁止宣称必须使用 blocked。", row=index, field="generation_policy"))

        identity = (sku, attribute, value.casefold(), normalized["market_scope"])
        if identity in seen:
            issues.append(_issue("warning", "duplicate_fact", "重复事实已忽略。", row=index))
            continue
        seen.add(identity)
        digest = hashlib.sha256(f"{project_id}|{sku}|{attribute}|{value}|{normalized['source']}".encode()).hexdigest()[:16]
        facts.append({"fact_id": f"BETA-{digest.upper()}", "project_id": project_id, "row": index, **normalized, "markets": markets, "status": "unconfirmed"})
    return facts, issues


def _dataset_issues(facts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    by_sku: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for fact in facts:
        if fact["sku"]:
            by_sku[fact["sku"]].append(fact)
    if not by_sku:
        return [_issue("error", "empty_import", "文件没有可读取的事实。")]
    if len(by_sku) > 3:
        issues.append(_issue("error", "too_many_skus", "单个项目最多导入 3 个 SKU。"))
    for sku, sku_facts in by_sku.items():
        attributes = {fact["attribute"] for fact in sku_facts}
        for missing in sorted(REQUIRED_ATTRIBUTES - attributes):
            issues.append(_issue("error", "missing_required_attribute", f"{sku} 缺少必填事实：{missing}。", field=missing))
        grouped: dict[str, set[str]] = defaultdict(set)
        for fact in sku_facts:
            if fact["attribute"] in SINGLE_VALUE_ATTRIBUTES:
                grouped[fact["attribute"]].add(fact["value"].casefold())
        for attribute, values in grouped.items():
            if len(values) > 1:
                issues.append(_issue("error", "conflicting_values", f"{sku} 的 {attribute} 存在冲突值。", field=attribute))
    return issues


def parse_beta_upload(filename: str, data: bytes, *, project_id: str) -> dict[str, Any]:
    if not PROJECT_PATTERN.fullmatch(project_id):
        raise BetaImportError("project_id 格式无效。")
    if not data:
        raise BetaImportError("上传文件为空。")
    if len(data) > MAX_UPLOAD_BYTES:
        raise BetaImportError("上传文件超过 2 MB。")
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        rows = _rows_from_csv(data)
    elif extension == "json":
        rows = _rows_from_json(data)
    elif extension == "xlsx":
        rows = _rows_from_xlsx(data)
    else:
        raise BetaImportError("只支持 CSV、JSON 和 XLSX。")
    facts, issues = _normalize_rows(rows, project_id)
    issues.extend(_dataset_issues(facts))
    errors = [issue for issue in issues if issue["severity"] == "error"]
    return {
        "project_id": project_id,
        "filename": filename.rsplit("/", 1)[-1].rsplit("\\", 1)[-1],
        "content_digest": hashlib.sha256(data).hexdigest(),
        "parsed_at": datetime.now(UTC).isoformat(),
        "facts": facts,
        "issues": issues,
        "summary": {
            "sku_count": len({fact["sku"] for fact in facts}),
            "fact_count": len(facts),
            "error_count": len(errors),
            "warning_count": sum(issue["severity"] == "warning" for issue in issues),
        },
        "ready_for_confirmation": not errors,
        "generation_enabled": False,
    }


def confirm_beta_import(preview: dict[str, Any], *, confirmed_by: str) -> dict[str, Any]:
    if not preview.get("ready_for_confirmation"):
        raise BetaImportError("仍有阻断问题，不能确认导入。")
    if not confirmed_by or len(confirmed_by) > 128:
        raise BetaImportError("确认人标识无效。")
    confirmed_at = datetime.now(UTC).isoformat()
    result = json.loads(json.dumps(preview, ensure_ascii=False))
    for fact in result["facts"]:
        fact["status"] = "confirmed"
        fact["confirmed_by"] = confirmed_by
        fact["confirmed_at"] = confirmed_at
    result["confirmation"] = {"confirmed_by": confirmed_by, "confirmed_at": confirmed_at}
    result["generation_enabled"] = True
    return result


class BetaProjectStore:
    """In-memory authorization model for tests and local sessions only."""

    def __init__(self) -> None:
        self._owners: dict[str, str] = {}
        self._projects: dict[str, dict[str, Any]] = {}

    def create(self, *, project_id: str, owner_id: str) -> None:
        if project_id in self._owners or not owner_id:
            raise BetaImportError("项目已存在或 owner_id 无效。")
        self._owners[project_id] = owner_id

    def save(self, *, project_id: str, actor_id: str, payload: dict[str, Any]) -> None:
        self._authorize(project_id, actor_id)
        self._projects[project_id] = json.loads(json.dumps(payload, ensure_ascii=False))

    def get(self, *, project_id: str, actor_id: str) -> dict[str, Any] | None:
        self._authorize(project_id, actor_id)
        payload = self._projects.get(project_id)
        return json.loads(json.dumps(payload, ensure_ascii=False)) if payload else None

    def clear(self, *, project_id: str, actor_id: str) -> dict[str, str]:
        self._authorize(project_id, actor_id)
        self._projects.pop(project_id, None)
        return {"project_id": project_id, "deleted_at": datetime.now(UTC).isoformat(), "status": "deleted"}

    def _authorize(self, project_id: str, actor_id: str) -> None:
        if self._owners.get(project_id) != actor_id:
            raise PermissionError("Project access denied.")
